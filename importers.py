import dateparser
import datetime

from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from .event_dict import EVENT_DICT
from .models import *

def get_unit_for_event(event_name):
    unit = 'inches'
    if 'meter' in event_name.lower():
        unit = 'seconds'
    elif 'smr' in event_name.lower():
        unit = 'seconds'
    elif 'hurdle' in event_name.lower():
        unit = 'seconds'
    elif 'hurdle' in event_name.lower():
        unit = 'seconds'
    elif 'yard' in event_name.lower():
        unit = 'seconds'
    elif 'mile' in event_name.lower():
        unit = 'seconds'
    elif 'medley' in event_name.lower():
        unit = 'seconds'
    return unit




def import_performances(data_file, team, season, gender):
    wb = load_workbook(data_file)
    sheet = wb.active

    meets = {}
    for meet in Meet.objects.filter(season=season, team=team):
        key = f"{meet.description}--{meet.season.name}"
        meets[key] = meet

    with transaction.atomic():
        headers = []
        for header in sheet[1]:
            headers.append(header.value.lower())

        total = 0
        users = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, row))

            if not (row['last name'] or row['first name']):
                continue

            first_name = row['first name'].strip().capitalize()
            last_name = row['last name'].strip().capitalize()
            username = f"{first_name.lower()}.{last_name.lower()}"

            try:
                user = User.objects.get(
                    first_name=first_name,
                    last_name=last_name
                )
            except User.DoesNotExist:
                print(f"Creating user {username}")
                user = User(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    gender=gender
                )
                user.save()
            users.add(user)

            event_name = row['event']
            try:
                event_name = str(int(event_name))
            except:
                pass
            event_name = str(event_name).strip()

            if event_name in EVENT_DICT:
                if EVENT_DICT[event_name] != '':
                    event_name = EVENT_DICT[event_name]
            else:
                raise Exception(f"Unknown event {event_name}")

            unit = get_unit_for_event(event_name)
            try:
                event = Event.objects.get(name=event_name)
            except Event.DoesNotExist:
                print(f"Creating event {event_name}")
                event = Event(
                    name=event_name,
                    unit=unit
                )
                event.save()

            if 'meet' in row:
                meet_name = row['meet']
                meet_date = row['date'].date()
            elif 'opponent' in row:
                meet_name = row['opponent']
                meet_date = row['date']
                if isinstance(meet_date, str):
                    meet_date = dateparser.parse(meet_date)
                else:
                    meet_date = meet_date.date()
            else:
                meet_name = row['date']
                date_str = meet_name.split(' ')[0]
                date_str += '/2019'
                meet_date = datetime.datetime.strptime(date_str, "%m/%d/%Y").date()

            meet_name = meet_name.strip()
            meet_name = meet_name.replace("Cetnral", "Central")

            key = f"{meet_name}--{season.name}"
            meet = meets.get(key)
            if not meet:
                print(f"Creating {meet_name} for {team.name}")
                meet = Meet(
                    description=meet_name,
                    season=season,
                    date=meet_date,
                    team=team,
                )
                meet.save()
                meets[key] = meet

            performance = row['performance']
            if isinstance(performance, str):
                # Fix common mistakes:
                performance = performance.replace('..', '.')

                if "-" in performance:
                    feet, inches = performance.split("-")
                    feet = float(feet)
                    inches = float(inches)
                    performance = (12.0 * feet) + inches
                elif ':' in performance:
                    performance = performance.replace(',', '.')
                    if '.' not in performance:
                        performance += '.0'

                    try:
                        pt = datetime.datetime.strptime(performance,'%M:%S.%f')
                    except:
                        print(f"Bad peformance {performance} for {user.username}")
                        raise
                    performance = pt.second + pt.minute*60.0 + pt.hour*3600.0 + (pt.microsecond/1000000.0)

            try:
                performance = float(performance)
            except:
                print(f"*** Bad peformance {performance} for {user.username}")
                continue

            try:
                result = Result.objects.get(
                    meet=meet,
                    event=event,
                    athlete=user,
                    result=performance
                )
            except Result.DoesNotExist:
                if 'fat/ht/na' in row:
                    method = row['fat/ht/na']
                else:
                    method = row['fat / hand']

                if method == 'HT':
                    method = 'Hand'

                result = Result(
                    meet=meet,
                    event=event,
                    athlete=user,
                    result=performance,
                    method=method
                )
                result.save()
                total += 1

        # Recalc
        for user in users:
            calculate_result_stats(user)

        #raise Exception("Data Test.")


    print(f"{total} rows processed") 



def import_qualifying(data_file, season):
    wb = load_workbook(data_file)
    sheet = wb.active
    with transaction.atomic():
        headers = []
        for header in sheet[1]:
            headers.append(header.value.lower())

        total = 0
        users = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, row))

            description = row['description']
            if not description:
                continue
            description = description.strip()
            gender = row['gender'].strip()

            event_name = row['event'].strip()
            if event_name in EVENT_DICT:
                if EVENT_DICT[event_name] != '':
                    event_name = EVENT_DICT[event_name]

            try:
                event = Event.objects.get(name=event_name)
            except Event.DoesNotExist:
                print(f"Creating {event_name}")
                unit = get_unit_for_event(event_name)
                event = Event(
                    name=event_name,
                    unit=unit
                )
                event.save()


            performance = row['performance']
            if performance == 'NA':
                continue
            if isinstance(performance, str):
                if "-" in performance:
                    feet, inches = performance.split("-")
                    feet = float(feet)
                    inches = float(inches)
                    performance = (12.0 * feet) + inches
                elif ':' in performance:
                    try:
                        pt = datetime.datetime.strptime(performance,'%M:%S.%f')
                    except:
                        import pdb; pdb.set_trace()
                        print(f"Skipping {performance}")
                        continue
                    performance = pt.second + pt.minute*60.0 + pt.hour*3600.0 + (pt.microsecond/1000000.0)
            elif isinstance(performance, datetime.time):
                    performance = (performance.minute * 60) + performance.second + (performance.microsecond / 1000000.0)
            try:
                performance = float(performance)
            except:
                import pdb; pdb.set_trace()
                print(f"Skipping {performance}")
                continue

            try:
                qt = QualifyingLevel.objects.get(
                    description=description,
                    event=event,
                    season=season,
                    gender=gender
                )
            except QualifyingLevel.DoesNotExist:
                print(f"Creating {event.name} for {description} for {season.name}")
                qt = QualifyingLevel(
                    description=description,
                    event=event,
                    season=season,
                    gender=gender
                )
            qt.value = performance
            qt.save()
            
